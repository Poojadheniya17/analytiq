# backend/routers/export.py
# Analytiq — Professional Export Router

import sys, os, json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from core.pdf_export import generate_pdf
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

router = APIRouter()

BASE_DATA = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "users")
)

# ── Style constants ───────────────────────────────────────────
BLUE_DARK  = "0F172A"
BLUE_MID   = "2563EB"
BLUE_LIGHT = "DBEAFE"
GRAY_LIGHT = "F8FAFC"
GRAY_MID   = "E2E8F0"
WHITE      = "FFFFFF"
RED        = "EF4444"
RED_LIGHT  = "FEE2E2"
GREEN      = "10B981"
GREEN_LIGHT= "D1FAE5"
ORANGE     = "F97316"
ORANGE_LIGHT="FFEDD5"

def hdr(bold=True, size=10, color=WHITE, bg=BLUE_DARK):
    f = Font(bold=bold, size=size, color=color, name="Calibri")
    p = PatternFill("solid", fgColor=bg)
    a = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return f, p, a

def cell_style(ws, cell_ref, value, bold=False, size=10, color="000000",
               bg=None, align="left", border=False, wrap=False, number_format=None):
    cell = ws[cell_ref] if isinstance(cell_ref, str) else cell_ref
    cell.value = value
    cell.font  = Font(bold=bold, size=size, color=color, name="Calibri")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if border:
        thin = Side(style="thin", color=GRAY_MID)
        cell.border = Border(bottom=thin)
    if number_format:
        cell.number_format = number_format
    return cell

def thin_border():
    t = Side(style="thin", color=GRAY_MID)
    return Border(top=t, left=t, right=t, bottom=t)

def bottom_border():
    t = Side(style="thin", color=GRAY_MID)
    return Border(bottom=t)

def get_path(user_id: int, client_name: str) -> str:
    return os.path.join(BASE_DATA, str(user_id), client_name.lower().replace(" ", "_"))


class ExportRequest(BaseModel):
    user_id:     int
    client_name: str
    domain:      str


@router.post("/pdf")
def export_pdf(req: ExportRequest):
    client_path    = get_path(req.user_id, req.client_name)
    insights_path  = os.path.join(client_path, "insights.json")
    metrics_path   = os.path.join(client_path, "model_metrics.json")
    narrative_path = os.path.join(client_path, "narrative.txt")
    charts_dir     = os.path.join(client_path, "charts")
    output_path    = os.path.join(client_path, "report.pdf")

    if not os.path.exists(insights_path):
        raise HTTPException(status_code=404, detail="Run insights first.")

    pdf_path = generate_pdf(
        client_name    = req.client_name,
        domain         = req.domain,
        insights_path  = insights_path,
        metrics_path   = metrics_path   if os.path.exists(metrics_path)   else None,
        narrative_path = narrative_path if os.path.exists(narrative_path) else None,
        charts_dir     = charts_dir     if os.path.exists(charts_dir)     else None,
        output_path    = output_path
    )
    return FileResponse(pdf_path, media_type="application/pdf",
        filename=f"Analytiq_{req.client_name.replace(' ','_')}_Report.pdf")


@router.post("/excel")
def export_excel(req: ExportRequest):
    client_path    = get_path(req.user_id, req.client_name)
    insights_path  = os.path.join(client_path, "insights.json")
    metrics_path   = os.path.join(client_path, "model_metrics.json")
    at_risk_path   = os.path.join(client_path, "at_risk.csv")
    charts_dir     = os.path.join(client_path, "charts")
    recs_path      = os.path.join(client_path, "recommendations.json")
    full_path      = os.path.join(client_path, "model_full_results.json")
    xlsx_path      = os.path.join(client_path, "dashboard.xlsx")

    if not os.path.exists(insights_path):
        raise HTTPException(status_code=404, detail="Run insights first.")

    with open(insights_path) as f:
        ins = json.load(f)

    kpis     = ins.get("kpis", {})
    segments = ins.get("segment_insights", [])
    anomalies= ins.get("anomalies", [])

    metrics      = {}
    full_results = {}
    recommendations = []

    if os.path.exists(metrics_path):
        with open(metrics_path) as f: metrics = json.load(f)
    if os.path.exists(full_path):
        with open(full_path) as f:   full_results = json.load(f)
    if os.path.exists(recs_path):
        with open(recs_path) as f:   recommendations = json.load(f)

    now = datetime.now().strftime("%B %d, %Y")
    wb  = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.row_dimensions[1].height = 60

    # Header banner
    ws.merge_cells("A1:E1")
    h = ws["A1"]
    h.value = f"  Analytiq — {req.client_name}"
    h.font  = Font(bold=True, size=18, color=WHITE, name="Calibri")
    h.fill  = PatternFill("solid", fgColor=BLUE_MID)
    h.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = f"  {req.domain}  ·  Analytics Report  ·  {now}"
    sub.font  = Font(size=9, color=BLUE_LIGHT, name="Calibri")
    sub.fill  = PatternFill("solid", fgColor=BLUE_DARK)
    sub.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[3].height = 15

    # KPI cards row
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 35
    ws.row_dimensions[6].height = 15

    kpi_list    = list(kpis.items())[:4]
    kpi_colors  = [BLUE_MID, RED, GREEN, ORANGE]
    kpi_light   = [BLUE_LIGHT, RED_LIGHT, GREEN_LIGHT, ORANGE_LIGHT]
    kpi_cols    = ["B", "C", "D", "E"]

    for i, (k, v) in enumerate(kpi_list):
        col = kpi_cols[i]
        # Label
        lbl = ws[f"{col}4"]
        lbl.value     = k.replace("_", " ").replace("%","").upper().strip()
        lbl.font      = Font(bold=True, size=8, color=kpi_colors[i], name="Calibri")
        lbl.fill      = PatternFill("solid", fgColor=kpi_light[i])
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        # Value
        val_cell = ws[f"{col}5"]
        display  = f"{v}%" if "rate" in k or "%" in k else (f"{int(v):,}" if isinstance(v,(int,float)) else str(v))
        val_cell.value     = display
        val_cell.font      = Font(bold=True, size=20, color=kpi_colors[i], name="Calibri")
        val_cell.fill      = PatternFill("solid", fgColor=kpi_light[i])
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Key findings
    ws.row_dimensions[7].height = 20
    ws.merge_cells("B7:E7")
    ws["B7"].value     = "KEY FINDINGS"
    ws["B7"].font      = Font(bold=True, size=10, color=BLUE_MID, name="Calibri")
    ws["B7"].alignment = Alignment(horizontal="left", vertical="center")

    target_col = ins.get("target_col", "target").replace("_"," ").title()
    target_rate = kpis.get("target_rate_%", 0)
    total       = kpis.get("total_records", 0)
    total_pos   = kpis.get("total_positive", 0)

    findings = [
        f"Overall {target_col} rate: {target_rate}% — {int(total_pos):,} of {int(total):,} records affected.",
    ]
    if segments:
        s = segments[0]
        findings.append(f"Highest risk segment: '{s.get('highest_churn_value','N/A')}' in {s.get('segment','').replace('_',' ')} at {s.get('highest_churn_rate_%',0)}%.")
    if metrics:
        auc = metrics.get("auc_roc") or (metrics.get("metrics") or {}).get("auc_roc")
        if auc: findings.append(f"Predictive model (XGBoost) achieved AUC-ROC of {auc} — strong performance.")
    if recommendations:
        findings.append(f"{len(recommendations)} data-driven recommendations generated for immediate action.")

    for i, finding in enumerate(findings):
        row = 8 + i
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"B{row}:E{row}")
        c = ws[f"B{row}"]
        c.value     = f"  →  {finding}"
        c.font      = Font(size=9, color="000000" if i%2==0 else "334155", name="Calibri")
        c.fill      = PatternFill("solid", fgColor=GRAY_LIGHT if i%2==0 else WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Recommendations section
    start_row = 8 + len(findings) + 2
    ws.row_dimensions[start_row].height = 20
    ws.merge_cells(f"B{start_row}:E{start_row}")
    ws[f"B{start_row}"].value     = "STRATEGIC RECOMMENDATIONS"
    ws[f"B{start_row}"].font      = Font(bold=True, size=10, color=BLUE_MID, name="Calibri")
    ws[f"B{start_row}"].alignment = Alignment(horizontal="left", vertical="center")

    for i, rec in enumerate(recommendations[:3]):
        r = start_row + 1 + i * 3
        ws.row_dimensions[r].height   = 15
        ws.row_dimensions[r+1].height = 30
        ws.row_dimensions[r+2].height = 8

        ws.merge_cells(f"B{r}:E{r}")
        title_cell = ws[f"B{r}"]
        title_cell.value     = f"  {i+1}.  {rec.get('title','')}"
        title_cell.font      = Font(bold=True, size=9, color=WHITE, name="Calibri")
        title_cell.fill      = PatternFill("solid", fgColor=[BLUE_MID, "059669", "EA580C"][i])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws[f"B{r+1}"].value     = f"Action: {rec.get('action','')}"
        ws[f"B{r+1}"].font      = Font(size=8.5, name="Calibri")
        ws[f"B{r+1}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"B{r+1}"].fill      = PatternFill("solid", fgColor=GRAY_LIGHT)
        ws.merge_cells(f"B{r+1}:C{r+1}")

        ws[f"D{r+1}"].value     = f"Impact: {rec.get('impact','')}"
        ws[f"D{r+1}"].font      = Font(size=8.5, name="Calibri", color="166534")
        ws[f"D{r+1}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[f"D{r+1}"].fill      = PatternFill("solid", fgColor=GRAY_LIGHT)
        ws.merge_cells(f"D{r+1}:E{r+1}")

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — KPI DASHBOARD
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("KPI Dashboard")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 35

    ws2.row_dimensions[1].height = 35
    ws2.merge_cells("A1:C1")
    ws2["A1"].value     = f"KPI Dashboard  ·  {req.client_name}"
    ws2["A1"].font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
    ws2["A1"].fill      = PatternFill("solid", fgColor=BLUE_MID)
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws2.row_dimensions[2].height = 20
    for col, header in zip(["A","B","C"], ["Metric", "Value", "Commentary"]):
        c = ws2[f"{col}2"]
        c.value = header
        c.font  = Font(bold=True, size=9, color=WHITE, name="Calibri")
        c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")

    metric_comments = {
        "total_records":  "Total number of records in the dataset",
        "target_rate_%":  "Percentage of records classified as positive target",
        "total_positive": "Count of positive target records requiring attention",
        "total_negative": "Count of negative/healthy records",
        "avg_monthly_charges": "Average monthly revenue per customer",
        "total_features": "Number of features used in the analysis",
    }

    for i, (k, v) in enumerate(kpis.items()):
        row = 3 + i
        ws2.row_dimensions[row].height = 18
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE

        ws2[f"A{row}"].value     = k.replace("_"," ").title()
        ws2[f"A{row}"].font      = Font(size=9, name="Calibri")
        ws2[f"A{row}"].fill      = PatternFill("solid", fgColor=bg)
        ws2[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws2[f"A{row}"].border    = bottom_border()

        display = f"{v}%" if "rate" in k or "%" in k else (f"{int(v):,}" if isinstance(v,(int,float)) else str(v))
        ws2[f"B{row}"].value     = display
        ws2[f"B{row}"].font      = Font(bold=True, size=9, color=BLUE_MID, name="Calibri")
        ws2[f"B{row}"].fill      = PatternFill("solid", fgColor=bg)
        ws2[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws2[f"B{row}"].border    = bottom_border()

        comment = metric_comments.get(k, "Derived from dataset analysis")
        ws2[f"C{row}"].value     = comment
        ws2[f"C{row}"].font      = Font(size=8.5, color="64748B", name="Calibri")
        ws2[f"C{row}"].fill      = PatternFill("solid", fgColor=bg)
        ws2[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws2[f"C{row}"].border    = bottom_border()

    # ══════════════════════════════════════════════════════════
    # SHEET 3 — AT-RISK RECORDS (color coded)
    # ══════════════════════════════════════════════════════════
    if os.path.exists(at_risk_path):
        try:
            df_risk = pd.read_csv(at_risk_path).head(50)
            ws3 = wb.create_sheet("At-Risk Records")
            ws3.sheet_view.showGridLines = False

            ws3.row_dimensions[1].height = 35
            ws3.merge_cells(f"A1:{get_column_letter(min(len(df_risk.columns),8))}1")
            ws3["A1"].value     = f"At-Risk Records  ·  {req.client_name}  ·  Top {len(df_risk)} highest risk"
            ws3["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
            ws3["A1"].fill      = PatternFill("solid", fgColor=RED)
            ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")

            ws3.row_dimensions[2].height = 20
            cols = list(df_risk.columns)[:8]
            for j, col in enumerate(cols):
                c = ws3.cell(row=2, column=j+1)
                c.value = col.replace("_"," ").upper()[:15]
                c.font  = Font(bold=True, size=8.5, color=WHITE, name="Calibri")
                c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
                c.alignment = Alignment(horizontal="center", vertical="center")
                ws3.column_dimensions[get_column_letter(j+1)].width = 18

            for i, (_, row) in enumerate(df_risk.iterrows()):
                excel_row = i + 3
                ws3.row_dimensions[excel_row].height = 16
                bg = GRAY_LIGHT if i % 2 == 0 else WHITE

                for j, col in enumerate(cols):
                    val  = row[col]
                    cell = ws3.cell(row=excel_row, column=j+1)

                    # Color code risk probability column
                    if "risk" in col.lower() or "prob" in col.lower():
                        try:
                            risk_val = float(val)
                            if risk_val > 70:
                                cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
                                cell.font = Font(bold=True, size=8.5, color=RED, name="Calibri")
                            elif risk_val > 40:
                                cell.fill = PatternFill("solid", fgColor=ORANGE_LIGHT)
                                cell.font = Font(bold=True, size=8.5, color=ORANGE, name="Calibri")
                            else:
                                cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
                                cell.font = Font(bold=True, size=8.5, color=GREEN, name="Calibri")
                        except:
                            cell.fill = PatternFill("solid", fgColor=bg)
                            cell.font = Font(size=8.5, name="Calibri")
                    else:
                        cell.fill = PatternFill("solid", fgColor=bg)
                        cell.font = Font(size=8.5, name="Calibri")

                    cell.value     = str(val)[:20] if pd.notna(val) else ""
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = bottom_border()
        except Exception as e:
            print(f"At-risk sheet error: {e}")

    # ══════════════════════════════════════════════════════════
    # SHEET 4 — SEGMENT ANALYSIS
    # ══════════════════════════════════════════════════════════
    if segments:
        ws4 = wb.create_sheet("Segment Analysis")
        ws4.sheet_view.showGridLines = False
        ws4.column_dimensions["A"].width = 3
        ws4.column_dimensions["B"].width = 25
        ws4.column_dimensions["C"].width = 25
        ws4.column_dimensions["D"].width = 15
        ws4.column_dimensions["E"].width = 20

        ws4.row_dimensions[1].height = 35
        ws4.merge_cells("A1:E1")
        ws4["A1"].value     = f"Segment Analysis  ·  {req.client_name}"
        ws4["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
        ws4["A1"].fill      = PatternFill("solid", fgColor=BLUE_MID)
        ws4["A1"].alignment = Alignment(horizontal="left", vertical="center")

        current_row = 2
        seg_colors  = [BLUE_MID, "059669", "EA580C", "7C3AED"]

        for si, seg in enumerate(segments[:4]):
            seg_name  = seg.get("segment","").replace("_"," ").title()
            breakdown = seg.get("full_breakdown", [])

            ws4.row_dimensions[current_row].height = 22
            ws4.merge_cells(f"B{current_row}:E{current_row}")
            ws4[f"B{current_row}"].value     = f"  {seg_name}"
            ws4[f"B{current_row}"].font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
            ws4[f"B{current_row}"].fill      = PatternFill("solid", fgColor=seg_colors[si % 4])
            ws4[f"B{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            ws4.row_dimensions[current_row].height = 16
            for col, header in zip(["B","C","D","E"], ["Segment Value","Count","Churn Rate %","Risk Level"]):
                c = ws4[f"{col}{current_row}"]
                c.value = header
                c.font  = Font(bold=True, size=8, color=WHITE, name="Calibri")
                c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
                c.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

            for item in breakdown[:8]:
                ws4.row_dimensions[current_row].height = 16
                vals = list(item.values())
                seg_val   = str(vals[0])[:25] if len(vals) > 0 else ""
                count_val = str(vals[1])[:15] if len(vals) > 1 else ""
                rate_val  = float(vals[2]) if len(vals) > 2 else 0
                bg = GRAY_LIGHT if current_row % 2 == 0 else WHITE

                ws4[f"B{current_row}"].value = seg_val
                ws4[f"B{current_row}"].font  = Font(size=8.5, name="Calibri")
                ws4[f"B{current_row}"].fill  = PatternFill("solid", fgColor=bg)
                ws4[f"B{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
                ws4[f"B{current_row}"].border = bottom_border()

                ws4[f"C{current_row}"].value = count_val
                ws4[f"C{current_row}"].font  = Font(size=8.5, name="Calibri")
                ws4[f"C{current_row}"].fill  = PatternFill("solid", fgColor=bg)
                ws4[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws4[f"C{current_row}"].border = bottom_border()

                # Color code rate
                rate_color = RED if rate_val > 50 else ORANGE if rate_val > 25 else GREEN
                rate_bg    = RED_LIGHT if rate_val > 50 else ORANGE_LIGHT if rate_val > 25 else GREEN_LIGHT
                ws4[f"D{current_row}"].value = f"{rate_val}%"
                ws4[f"D{current_row}"].font  = Font(bold=True, size=8.5, color=rate_color, name="Calibri")
                ws4[f"D{current_row}"].fill  = PatternFill("solid", fgColor=rate_bg)
                ws4[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws4[f"D{current_row}"].border = bottom_border()

                risk_label = "HIGH" if rate_val > 50 else "MEDIUM" if rate_val > 25 else "LOW"
                ws4[f"E{current_row}"].value = risk_label
                ws4[f"E{current_row}"].font  = Font(bold=True, size=8, color=rate_color, name="Calibri")
                ws4[f"E{current_row}"].fill  = PatternFill("solid", fgColor=rate_bg)
                ws4[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws4[f"E{current_row}"].border = bottom_border()

                current_row += 1
            current_row += 1

    # ══════════════════════════════════════════════════════════
    # SHEET 5 — MODEL PERFORMANCE
    # ══════════════════════════════════════════════════════════
    if metrics:
        ws5 = wb.create_sheet("Model Performance")
        ws5.sheet_view.showGridLines = False
        ws5.column_dimensions["A"].width = 3
        ws5.column_dimensions["B"].width = 25
        ws5.column_dimensions["C"].width = 15
        ws5.column_dimensions["D"].width = 40
        ws5.column_dimensions["E"].width = 20

        ws5.row_dimensions[1].height = 35
        ws5.merge_cells("A1:E1")
        ws5["A1"].value     = f"Model Performance  ·  {req.client_name}"
        ws5["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
        ws5["A1"].fill      = PatternFill("solid", fgColor=BLUE_MID)
        ws5["A1"].alignment = Alignment(horizontal="left", vertical="center")

        # Model info
        best_model   = full_results.get("best_model") or metrics.get("best_model","XGBoost")
        problem_type = full_results.get("problem_type") or metrics.get("problem_type","classification")

        info_items = [
            ("Best Model",    best_model),
            ("Problem Type",  problem_type.replace("_"," ").title()),
            ("Target Column", full_results.get("target_column") or metrics.get("target_column","N/A")),
            ("Features Used", str(full_results.get("feature_count","N/A"))),
            ("Training Rows", f"{full_results.get('training_rows',0):,}" if full_results.get("training_rows") else "N/A"),
        ]

        for i, (label, value) in enumerate(info_items):
            row = 2 + i
            ws5.row_dimensions[row].height = 18
            bg = GRAY_LIGHT if i % 2 == 0 else WHITE
            ws5[f"B{row}"].value = label
            ws5[f"B{row}"].font  = Font(size=9, name="Calibri", color="64748B")
            ws5[f"B{row}"].fill  = PatternFill("solid", fgColor=bg)
            ws5[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws5[f"B{row}"].border = bottom_border()
            ws5.merge_cells(f"C{row}:E{row}")
            ws5[f"C{row}"].value = value
            ws5[f"C{row}"].font  = Font(bold=True, size=9, name="Calibri")
            ws5[f"C{row}"].fill  = PatternFill("solid", fgColor=bg)
            ws5[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws5[f"C{row}"].border = bottom_border()

        # Metrics
        metric_row = 8
        ws5.row_dimensions[metric_row].height = 22
        for col, header in zip(["B","C","D","E"], ["Metric","Score","What it means","Rating"]):
            c = ws5[f"{col}{metric_row}"]
            c.value = header
            c.font  = Font(bold=True, size=9, color=WHITE, name="Calibri")
            c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")

        explanations = {
            "auc_roc":   ("Overall model quality. > 0.85 is excellent.",     0.85),
            "precision": ("Of all predicted positives, % that were correct.", 0.80),
            "recall":    ("Of all actual positives, % that were found.",      0.75),
            "f1_score":  ("Harmonic mean of precision and recall.",           0.80),
            "accuracy":  ("Overall % of correct predictions.",                0.85),
            "rmse":      ("Root mean square error. Lower is better.",         None),
            "mae":       ("Mean absolute error. Lower is better.",            None),
            "r2_score":  ("% of variance explained by model.",                0.80),
        }

        actual_metrics = full_results.get("metrics") or metrics
        if isinstance(actual_metrics, dict):
            for i, (k, v) in enumerate(actual_metrics.items()):
                if not isinstance(v, (int, float)): continue
                row = metric_row + 1 + i
                ws5.row_dimensions[row].height = 18
                bg = GRAY_LIGHT if i % 2 == 0 else WHITE
                explanation, threshold = explanations.get(k, ("Performance metric", None))

                rating = ""
                if threshold:
                    rating = "Excellent" if v >= threshold else "Good" if v >= threshold*0.9 else "Fair"
                    r_color = GREEN if v >= threshold else ORANGE if v >= threshold*0.9 else RED
                    r_bg    = GREEN_LIGHT if v >= threshold else ORANGE_LIGHT if v >= threshold*0.9 else RED_LIGHT

                ws5[f"B{row}"].value = k.replace("_"," ").upper()
                ws5[f"B{row}"].font  = Font(size=8.5, name="Calibri")
                ws5[f"B{row}"].fill  = PatternFill("solid", fgColor=bg)
                ws5[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
                ws5[f"B{row}"].border = bottom_border()

                ws5[f"C{row}"].value = round(v, 4)
                ws5[f"C{row}"].font  = Font(bold=True, size=9, color=BLUE_MID, name="Calibri")
                ws5[f"C{row}"].fill  = PatternFill("solid", fgColor=bg)
                ws5[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws5[f"C{row}"].border = bottom_border()

                ws5[f"D{row}"].value = explanation
                ws5[f"D{row}"].font  = Font(size=8, color="64748B", name="Calibri")
                ws5[f"D{row}"].fill  = PatternFill("solid", fgColor=bg)
                ws5[f"D{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws5[f"D{row}"].border = bottom_border()

                if rating:
                    ws5[f"E{row}"].value = rating
                    ws5[f"E{row}"].font  = Font(bold=True, size=8, color=r_color, name="Calibri")
                    ws5[f"E{row}"].fill  = PatternFill("solid", fgColor=r_bg)
                    ws5[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")
                    ws5[f"E{row}"].border = bottom_border()

    # ══════════════════════════════════════════════════════════
    # SHEET 6 — VISUAL INSIGHTS (charts)
    # ══════════════════════════════════════════════════════════
    if os.path.exists(charts_dir):
        charts = [f for f in os.listdir(charts_dir) if f.endswith(".png")]
        if charts:
            ws6 = wb.create_sheet("Visual Insights")
            ws6.sheet_view.showGridLines = False
            ws6.row_dimensions[1].height = 35
            ws6.merge_cells("A1:N1")
            ws6["A1"].value     = f"Visual Insights  ·  {req.client_name}"
            ws6["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
            ws6["A1"].fill      = PatternFill("solid", fgColor=BLUE_MID)
            ws6["A1"].alignment = Alignment(horizontal="left", vertical="center")

            positions = ["A3","N3","A33","N33","A63","N63"]
            for i, chart_file in enumerate(charts[:6]):
                try:
                    img = XLImage(os.path.join(charts_dir, chart_file))
                    img.width  = 480
                    img.height = 320
                    ws6.add_image(img, positions[i])

                    # Chart label
                    label_col = "A" if i % 2 == 0 else "N"
                    label_row = 3 + (i // 2) * 30
                    ws6[f"{label_col}{label_row}"].value = chart_file.replace("_"," ").replace(".png","").title()
                    ws6[f"{label_col}{label_row}"].font  = Font(bold=True, size=9, color=BLUE_MID, name="Calibri")
                except Exception as e:
                    print(f"Chart {chart_file} error: {e}")

    # ══════════════════════════════════════════════════════════
    # Save
    # ══════════════════════════════════════════════════════════
    wb.save(xlsx_path)
    return FileResponse(xlsx_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"Analytiq_{req.client_name.replace(' ','_')}_Dashboard.xlsx")
