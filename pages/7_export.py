# pages/7_export.py
import streamlit as st
import pandas as pd
import os, sys, json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from auth.auth import require_login
from core.pdf_export import generate_pdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Analytiq · Export", page_icon="◈", layout="wide")
exec(open(os.path.join(os.path.dirname(__file__), "_shared_css.py")).read())

user = require_login()

if "active_client" not in st.session_state:
    st.switch_page("pages/1_workspace.py")
    st.stop()

client      = st.session_state.active_client
client_path = st.session_state.active_client_path

with st.sidebar:
    st.markdown("<div style='padding:8px 0 16px'><div style='font-family:Fraunces,serif;font-size:20px;font-weight:600'>Analytiq</div></div>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-family:DM Mono,monospace;font-size:10px;color:#9CA3AF;line-height:2;margin-bottom:12px'>CLIENT<br><span style='font-size:13px;font-family:DM Sans,sans-serif;font-weight:500'>{client['name']}</span></div>", unsafe_allow_html=True)
    st.divider()
    if st.button("◈  Workspace"):      st.switch_page("pages/1_workspace.py")
    if st.button("↗  Upload & Clean"): st.switch_page("pages/2_upload.py")
    if st.button("◎  Insights"):       st.switch_page("pages/3_insights.py")
    if st.button("⬡  ML Model"):       st.switch_page("pages/4_model.py")
    if st.button("◌  Ask AI"):         st.switch_page("pages/5_ask_ai.py")
    if st.button("◻  Narrative"):      st.switch_page("pages/6_narrative.py")
    if st.button("↓  Export"):         st.switch_page("pages/7_export.py")
    st.divider()
    if st.button("Sign Out"):
        st.session_state.clear(); st.switch_page("app.py")

st.markdown("<h1>Export <em style='font-style:italic;color:#2563EB'>Report</em></h1>", unsafe_allow_html=True)
st.markdown(f"<p style='font-family:DM Mono,monospace;font-size:11px;color:#9CA3AF;margin-top:-8px;margin-bottom:24px'>Client-ready deliverables · {client['name']}</p>", unsafe_allow_html=True)

st.markdown("### Analysis Checklist")
checks = [
    (f"{client_path}/cleaned_data.csv",   "Data cleaned"),
    (f"{client_path}/insights.json",      "Insights generated"),
    (f"{client_path}/model_metrics.json", "ML model trained"),
    (f"{client_path}/narrative.txt",      "Executive narrative written"),
]
all_ready = True
for path, label in checks:
    if os.path.exists(path):
        st.success(f"✓  {label}")
    else:
        st.warning(f"○  {label} — not yet complete")
        all_ready = False

st.markdown("<br>", unsafe_allow_html=True)
st.markdown("### PDF Report")
st.markdown("<div style='background:#F7F8FA;border:1px solid #E2E8F0;border-radius:10px;padding:14px 18px;margin-bottom:16px;font-family:DM Mono,monospace;font-size:11px;color:#6B7280'>Generates a professional PDF with KPIs, segment analysis, model metrics, charts, and the AI executive narrative. Ready to send directly to a client.</div>", unsafe_allow_html=True)

if st.button("Generate PDF Report"):
    if not os.path.exists(f"{client_path}/insights.json"):
        st.error("Please complete the Insights step first.")
    else:
        with st.spinner("Building PDF..."):
            pdf_path = generate_pdf(
                client_name    = client["name"],
                domain         = client["domain"],
                insights_path  = f"{client_path}/insights.json",
                metrics_path   = f"{client_path}/model_metrics.json"  if os.path.exists(f"{client_path}/model_metrics.json") else None,
                narrative_path = f"{client_path}/narrative.txt"       if os.path.exists(f"{client_path}/narrative.txt")      else None,
                charts_dir     = f"{client_path}/charts"              if os.path.exists(f"{client_path}/charts")             else None,
                output_path    = f"{client_path}/report.pdf"
            )
        st.success("PDF generated!")
        with open(pdf_path, "rb") as f:
            st.download_button(
                label     = "Download PDF Report",
                data      = f,
                file_name = f"Analytiq_{client['name'].replace(' ','_')}_Report.pdf",
                mime      = "application/pdf"
            )

st.divider()
st.markdown("### Excel Dashboard")
st.markdown("<div style='background:#F7F8FA;border:1px solid #E2E8F0;border-radius:10px;padding:14px 18px;margin-bottom:16px;font-family:DM Mono,monospace;font-size:11px;color:#6B7280'>Exports KPIs, segment breakdown, at-risk records, and model metrics into a formatted Excel workbook.</div>", unsafe_allow_html=True)

if st.button("Generate Excel Dashboard"):
    if not os.path.exists(f"{client_path}/insights.json"):
        st.error("Please complete the Insights step first.")
    else:
        with st.spinner("Building Excel workbook..."):
            
            wb = openpyxl.Workbook()

            # ── KPI Sheet ─────────────────────────────────────
            ws = wb.active
            ws.title = "KPIs"

            with open(f"{client_path}/insights.json") as f:
                ins = json.load(f)

            header_fill = PatternFill("solid", fgColor="0F1117")
            header_font = Font(color="FFFFFF", bold=True, size=10)

            ws["A1"] = f"Analytiq Report - {client['name']} - {client['domain']}"
            ws["A1"].font = Font(bold=True, size=13)
            ws.merge_cells("A1:B1")
            ws.append([])
            ws.append(["Metric", "Value"])
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
            for k, v in ins["kpis"].items():
                ws.append([k.replace("_", " ").title(), v])
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 20

            # ── At-Risk Sheet ─────────────────────────────────
            at_risk_path = f"{client_path}/at_risk.csv"
            if os.path.exists(at_risk_path):
                ws2 = wb.create_sheet("At-Risk Records")
                df_risk = pd.read_csv(at_risk_path).head(25)
                ws2.append(list(df_risk.columns))
                for cell in ws2[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                for row in df_risk.itertuples(index=False):
                    ws2.append(list(row))
                for col in ws2.columns:
                    ws2.column_dimensions[col[0].column_letter].width = 18

            # ── Model Metrics Sheet ───────────────────────────
            if os.path.exists(f"{client_path}/model_metrics.json"):
                ws3 = wb.create_sheet("Model Metrics")
                with open(f"{client_path}/model_metrics.json") as f:
                    metrics = json.load(f)
                ws3.append(["Metric", "Score"])
                for cell in ws3[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                for k, v in metrics.items():
                    ws3.append([k.replace("_", " ").upper(), v])
                ws3.column_dimensions["A"].width = 20
                ws3.column_dimensions["B"].width = 15

            # ── Charts Sheet ──────────────────────────────────
            charts_dir = f"{client_path}/charts"
            if os.path.exists(charts_dir):
                chart_files = [f for f in os.listdir(charts_dir) if f.endswith(".png")]
                if chart_files:
                    ws4 = wb.create_sheet("Visual Insights")
                    ws4["A1"] = f"Visual Insights - {client['name']}"
                    ws4["A1"].font = Font(bold=True, size=13)

                    positions = ["A3", "M3", "A33", "M33", "A63", "M63"]
                    for i, chart_file in enumerate(chart_files[:6]):
                        try:
                            img_path = os.path.join(charts_dir, chart_file)
                            img = XLImage(img_path)
                            img.width  = 480
                            img.height = 320
                            ws4.add_image(img, positions[i])
                        except Exception as e:
                            ws4[f"A{3 + i * 30}"] = f"Chart unavailable: {chart_file}"

            xlsx_path = f"{client_path}/dashboard.xlsx"
            wb.save(xlsx_path)

        st.success("Excel dashboard generated!")
        with open(xlsx_path, "rb") as f:
            st.download_button(
                label     = "Download Excel Dashboard",
                data      = f,
                file_name = f"Analytiq_{client['name'].replace(' ','_')}_Dashboard.xlsx",
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )